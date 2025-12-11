from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import asyncio
import io
import csv
import os
import smtplib
from email.mime.text import MIMEText

from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    String,
    DateTime,
    Text,
    ForeignKey,
)
from sqlalchemy.orm import sessionmaker, declarative_base, Session, relationship

# ---------- DB SETUP ----------

SQLALCHEMY_DATABASE_URL = "sqlite:///./hr_email.db"

engine = create_engine(
    SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False}
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base = declarative_base()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------- MODELS ----------

class Recipient(Base):
    __tablename__ = "recipients"

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    email = Column(String, nullable=False, unique=True)
    # NEW: role for grouping – e.g., "Intern 3", "Manager"
    role = Column(String, nullable=True, index=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class Template(Base):
    __tablename__ = "templates"

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    subject = Column(String, nullable=False)
    body = Column(Text, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)


class EmailJob(Base):
    __tablename__ = "email_jobs"

    id = Column(Integer, primary_key=True, index=True)
    template_id = Column(Integer, ForeignKey("templates.id"), nullable=False)
    recipients_csv = Column(Text, nullable=False)  # comma-separated emails
    status = Column(String, default="scheduled")  # scheduled / processing / sent / failed
    scheduled_at = Column(DateTime, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    last_error = Column(Text, nullable=True)

    template = relationship("Template")


Base.metadata.create_all(bind=engine)

# ---------- SMTP CONFIG ----------

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", "hr@example.com")  # change for your org
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "true").lower() == "true"


# ---------- Pydantic SCHEMAS ----------

class RecipientCreate(BaseModel):
    name: str
    email: str
    role: Optional[str] = None  # NEW


class RecipientRead(BaseModel):
    id: int
    name: str
    email: str
    role: Optional[str] = None
    created_at: datetime

    class Config:
        orm_mode = True


class TemplateCreate(BaseModel):
    name: str
    subject: str
    body: str


class TemplateRead(BaseModel):
    id: int
    name: str
    subject: str
    body: str
    created_at: datetime

    class Config:
        orm_mode = True


class JobCreate(BaseModel):
    template_id: int

    # You now support:
    # - explicit recipients
    # - role-based selection
    # - send_to_all flag
    recipient_ids: List[int] = []
    roles: List[str] = []
    send_to_all: bool = False

    scheduled_at: Optional[str] = None  # ISO string from frontend; if None → send now


class JobRead(BaseModel):
    id: int
    template: TemplateRead
    recipients_csv: str
    status: str
    scheduled_at: datetime
    created_at: datetime
    last_error: Optional[str] = None

    class Config:
        orm_mode = True


# ---------- EMAIL SENDER (real SMTP + safe fallback) ----------

def send_email(to_email: str, subject: str, body: str):
    """
    If SMTP_* env vars are configured, send real email.
    Otherwise, just print to console (safe fallback).
    """
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS:
        # Fallback: console only
        print("=" * 60)
        print("[DRY-RUN] SENDING EMAIL (SMTP not configured)")
        print(f"TO: {to_email}")
        print(f"SUBJECT: {subject}")
        print("BODY:")
        print(body)
        print("=" * 60)
        return

    msg = MIMEText(body, "html")
    msg["Subject"] = subject
    msg["From"] = SMTP_FROM
    msg["To"] = to_email

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            if SMTP_USE_TLS:
                server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)
    except Exception as e:
        # Propagate error so job processor can mark job as failed
        raise RuntimeError(f"Failed to send email to {to_email}: {e}")


# ---------- BACKGROUND WORKER ----------

async def job_worker_loop():
    while True:
        db: Session = SessionLocal()
        try:
            now = datetime.utcnow()
            jobs = (
                db.query(EmailJob)
                .filter(EmailJob.status == "scheduled")
                .filter(EmailJob.scheduled_at <= now)
                .all()
            )
            for job in jobs:
                await process_job(db, job)
        finally:
            db.close()

        # check every 10 seconds
        await asyncio.sleep(10)


async def process_job(db: Session, job: EmailJob):
    job.status = "processing"
    db.commit()
    db.refresh(job)

    tmpl: Template = job.template
    emails = [e.strip() for e in job.recipients_csv.split(",") if e.strip()]
    errors = []

    for email in emails:
        try:
            # TODO: if you want personalization, do it here
            # e.g. subject/body.replace("{{email}}", email)
            send_email(email, tmpl.subject, tmpl.body)
        except Exception as e:
            errors.append(str(e))

    if errors:
        job.status = "failed"
        job.last_error = "; ".join(errors)
    else:
        job.status = "sent"
        job.last_error = None

    db.commit()


# ---------- FASTAPI APP ----------

app = FastAPI(title="HR Email Console", version="0.2.0")

# CORS for local frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # restrict in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup_event():
    # start background job worker
    asyncio.create_task(job_worker_loop())


# ----- Recipients -----


@app.get("/recipients", response_model=List[RecipientRead])
def list_recipients(db: Session = Depends(get_db)):
    return db.query(Recipient).order_by(Recipient.created_at.desc()).all()


@app.post("/recipients", response_model=RecipientRead)
def create_recipient(rec_in: RecipientCreate, db: Session = Depends(get_db)):
    # optional: check duplicate email
    existing = db.query(Recipient).filter(Recipient.email == rec_in.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")

    rec = Recipient(name=rec_in.name, email=rec_in.email, role=rec_in.role)
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return rec


@app.delete("/recipients/{recipient_id}")
def delete_recipient(recipient_id: int, db: Session = Depends(get_db)):
    rec = db.query(Recipient).get(recipient_id)
    if not rec:
        raise HTTPException(status_code=404, detail="Recipient not found")
    db.delete(rec)
    db.commit()
    return {"ok": True}


# ----- NEW: Bulk upload recipients from Excel/CSV -----


@app.post("/recipients/bulk_upload", response_model=List[RecipientRead])
async def bulk_upload_recipients(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Accepts .xlsx or .csv with columns: name, email, role
    - If email already exists, updates name/role.
    - Otherwise, creates new recipient.
    """
    filename = file.filename.lower()
    content = await file.read()

    rows = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    elif filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # assume first row is header
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # we expect "name", "email", "role"
        header_index = {h: i for i, h in enumerate(headers)}

        for r in ws.iter_rows(min_row=2):
            def get(col: str) -> str:
                idx = header_index.get(col)
                if idx is None:
                    return ""
                value = r[idx].value
                return str(value).strip() if value is not None else ""

            row = {
                "name": get("name"),
                "email": get("email"),
                "role": get("role"),
            }
            # skip empty lines
            if not row["email"]:
                continue
            rows.append(row)
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")

    created_or_updated: List[Recipient] = []

    for r in rows:
        name = r.get("name", "").strip()
        email = r.get("email", "").strip()
        role = r.get("role", "").strip() or None

        if not email or not name:
            # basic validation: skip incomplete lines
            continue

        existing = db.query(Recipient).filter(Recipient.email == email).first()
        if existing:
            # update name/role if changed
            existing.name = name
            existing.role = role
            created_or_updated.append(existing)
        else:
            rec = Recipient(name=name, email=email, role=role)
            db.add(rec)
            created_or_updated.append(rec)

    db.commit()
    # refresh instances to get IDs etc.
    for rec in created_or_updated:
        db.refresh(rec)

    return created_or_updated


# ----- NEW: Roles list (for role checkboxes in UI) -----


@app.get("/roles", response_model=List[str])
def list_roles(db: Session = Depends(get_db)):
    """
    Returns distinct non-null roles, e.g. ["Intern 3", "Intern 4", "Manager"]
    """
    rows = (
        db.query(Recipient.role)
        .filter(Recipient.role.isnot(None))
        .filter(Recipient.role != "")
        .distinct()
        .all()
    )
    return [r[0] for r in rows]


# ----- Templates -----


@app.get("/templates", response_model=List[TemplateRead])
def list_templates(db: Session = Depends(get_db)):
    return db.query(Template).order_by(Template.created_at.desc()).all()


@app.post("/templates", response_model=TemplateRead)
def create_template(tpl_in: TemplateCreate, db: Session = Depends(get_db)):
    tpl = Template(name=tpl_in.name, subject=tpl_in.subject, body=tpl_in.body)
    db.add(tpl)
    db.commit()
    db.refresh(tpl)
    return tpl


@app.delete("/templates/{template_id}")
def delete_template(template_id: int, db: Session = Depends(get_db)):
    tpl = db.query(Template).get(template_id)
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    db.delete(tpl)
    db.commit()
    return {"ok": True}


# ----- Jobs (Send / Schedule) -----


@app.get("/jobs", response_model=List[JobRead])
def list_jobs(db: Session = Depends(get_db)):
    jobs = db.query(EmailJob).order_by(EmailJob.created_at.desc()).all()
    return jobs


@app.post("/jobs", response_model=JobRead)
def create_job(job_in: JobCreate, db: Session = Depends(get_db)):
    tmpl = db.query(Template).get(job_in.template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")

    # Build recipient set
    recipients_q = db.query(Recipient)

    recipients_map = {}  # id -> Recipient

    # 1) send to all
    if job_in.send_to_all:
        for r in recipients_q.all():
            recipients_map[r.id] = r

    # 2) explicit IDs
    if job_in.recipient_ids:
        expl = (
            recipients_q.filter(Recipient.id.in_(job_in.recipient_ids)).all()
        )
        for r in expl:
            recipients_map[r.id] = r

    # 3) roles
    if job_in.roles:
        role_recs = (
            recipients_q.filter(Recipient.role.in_(job_in.roles)).all()
        )
        for r in role_recs:
            recipients_map[r.id] = r

    recipients = list(recipients_map.values())

    if not recipients:
        raise HTTPException(
            status_code=400,
            detail="No recipients selected. Use recipient_ids, roles, or send_to_all.",
        )

    emails_csv = ",".join(r.email for r in recipients)

    # parse scheduled time
    if job_in.scheduled_at:
        try:
            scheduled_at = datetime.fromisoformat(job_in.scheduled_at)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Invalid scheduled_at format (use ISO 8601)",
            )
    else:
        scheduled_at = datetime.utcnow()

    job = EmailJob(
        template_id=job_in.template_id,
        recipients_csv=emails_csv,
        scheduled_at=scheduled_at,
        status="scheduled",
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job
